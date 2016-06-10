import openpyxl


def normalize_header_name(header):
    if header:
        return header.strip().lower().replace(' ', '_').replace('/', '_')
    return None


def normalize(x):
    if not x:
        return ''
    return unicode(x)


def xlsx_to_dict_generator(f, tab_name):
    book = openpyxl.reader.excel.load_workbook(f, data_only=True)
    sheet = book.get_sheet_by_name(tab_name)

    rows = sheet.max_row
    cols = sheet.max_column

    headers = dict((i, sheet.cell(row=1, column=i).value) for i in xrange(1, cols + 1))

    # NOTE:book closes by the openpyxl handler, see http://stackoverflow.com/questions/11214908/closing-files-in-openpyxl

    return (
        {normalize_header_name(headers[column]): sheet.cell(row=row, column=column).value for column in
         xrange(1, cols + 1)}
        for row in
        xrange(2, rows + 1))
